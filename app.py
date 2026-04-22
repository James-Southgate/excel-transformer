import os
import io
import tempfile
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from flask import Flask, request, render_template_string, send_file

app = Flask(__name__)

# ------------------------------------------------------------
# Your EXACT transformation functions (copied from your script)
# ------------------------------------------------------------
def normalize(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return str(val).strip()

def load_with_computed_offer(file_path, sheet_name=None):
    # Step 1 — Load Excel with formula‑aware Offer column
    read_kwargs = {"dtype": str, "keep_default_na": False}
    if sheet_name:
        read_kwargs["sheet_name"] = sheet_name
    df = pd.read_excel(file_path, **read_kwargs)

    wb = load_workbook(file_path, data_only=False)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    try:
        offer_col_idx = headers.index('Offer') + 1
    except ValueError:
        return df

    offer_values = []
    for i in range(len(df)):
        excel_row = i + 2
        cell = ws.cell(row=excel_row, column=offer_col_idx)
        if cell.data_type == 'f':
            sale_str = df.loc[i, 'Sale Price']
            reg_str = df.loc[i, 'Reg Price']
            try:
                sale = float(''.join(c for c in str(sale_str) if c.isdigit() or c == '.'))
                reg = float(''.join(c for c in str(reg_str) if c.isdigit() or c == '.'))
                diff = reg - sale
                if diff == int(diff):
                    offer_text = f"Save ${int(diff)}"
                else:
                    offer_text = f"Save ${diff:.1f}".rstrip('0').rstrip('.')
            except (ValueError, TypeError):
                offer_text = None
        else:
            offer_text = cell.value if cell.value is not None else ''
        offer_values.append(normalize(offer_text))

    df['Offer'] = offer_values

    for col in ['Sale Price', 'Reg Price']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(r'[^0-9.]', '', regex=True)
            df[col] = df[col].str.strip()
    return df

def sort_by_department(df):
    return df.sort_values(by="Department", key=lambda s: s.str.lower(), kind="mergesort").reset_index(drop=True)

def insert_department_header_rows(df):
    rows = []
    previous_department = None
    for _, row in df.iterrows():
        current_department = normalize(row["Department"])
        if current_department != previous_department:
            separator = {col: "" for col in df.columns}
            separator["Department Headers"] = f"{current_department}.eps" if current_department else ".eps"
            rows.append(separator)
            previous_department = current_department
        data = row.to_dict()
        data["Department Headers"] = ""
        rows.append(data)
    return pd.DataFrame(rows, columns=df.columns)

def map_sale_type_eps(df):
    df["Sale Type"] = df["Sale Type"].apply(lambda v: f"{normalize(v)}.eps" if normalize(v) else "")
    return df

def process_offer(df):
    offer_idx = df.columns.get_loc("Offer")
    new_dollar = []
    new_offer_dollar = []
    new_offer_cents = []
    new_offer = []
    for val in df["Offer"]:
        original = normalize(val)
        if "$" in original:
            prefix, price = original.split("$", 1)
            price = price.strip()
            if price.startswith("0."):
                digits_after_decimal = price.split(".", 1)[1]
                new_offer.append(prefix)
                new_dollar.append("")
                new_offer_dollar.append(digits_after_decimal)
                new_offer_cents.append("¢")
            else:
                new_offer.append(prefix)
                new_dollar.append("$")
                if "." in price:
                    dollars, cents = price.split(".", 1)
                    new_offer_dollar.append(dollars)
                    new_offer_cents.append(cents)
                else:
                    new_offer_dollar.append(price)
                    new_offer_cents.append("")
        else:
            new_offer.append(original)
            new_dollar.append("")
            new_offer_dollar.append("")
            new_offer_cents.append("")
    df["Offer"] = new_offer
    df.insert(offer_idx + 1, "$", new_dollar)
    df.insert(offer_idx + 2, "Offer Dollar", new_offer_dollar)
    df.insert(offer_idx + 3, "Offer Cents", new_offer_cents)
    return df

def process_sale_price(df):
    sp_idx = df.columns.get_loc("Sale Price")
    new_sale_price = []
    new_sale_cents = []
    for val in df["Sale Price"]:
        s = normalize(val)
        if s == "":
            new_sale_price.append("")
            new_sale_cents.append("")
            continue
        if "." in s:
            dollars, cents = s.split(".", 1)
        else:
            dollars, cents = s, "00"
        if len(cents) == 1:
            cents = f"0{cents}"
        if dollars == "0":
            new_sale_price.append(cents)
            new_sale_cents.append("¢")
        else:
            new_sale_price.append(dollars)
            new_sale_cents.append(cents)
    df["Sale Price"] = new_sale_price
    df.insert(sp_idx + 1, "Sale Cents", new_sale_cents)
    return df

def process_reg_price(df):
    df["Reg Price"] = df["Reg Price"].apply(
        lambda v: f"Reg Price ${normalize(v)}.00 |" if "." not in normalize(v) and normalize(v) else
                  (f"Reg Price ${normalize(v)} |" if normalize(v) else "")
    )
    return df

def transform_in_memory(input_bytes):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
        tmp_in.write(input_bytes)
        tmp_in_path = tmp_in.name
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
        tmp_out_path = tmp_out.name
    try:
        df = load_with_computed_offer(tmp_in_path)
        df = sort_by_department(df)
        df.insert(0, "Department Headers", "")
        df = insert_department_header_rows(df)
        df = map_sale_type_eps(df)
        df = process_offer(df)
        df = process_sale_price(df)
        df = process_reg_price(df)
        df.to_excel(tmp_out_path, index=False, engine="openpyxl")
        with open(tmp_out_path, "rb") as f:
            return f.read()
    finally:
        os.unlink(tmp_in_path)
        os.unlink(tmp_out_path)

# ------------------------------------------------------------
# Flask routes
# ------------------------------------------------------------
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Cleanup Tool | Transform your data</title>
    <!-- Tailwind CSS + Font Awesome + SheetJS -->
    <script src="https://cdn.tailwindcss.com"></script>
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
    <script src="https://cdn.sheetjs.com/xlsx-0.20.2/package/dist/xlsx.full.min.js"></script>
    <style>
        /* Custom smooth transitions */
        .transition-all { transition: all 0.2s ease; }
        .preview-table { overflow-x: auto; max-height: 300px; }
        .preview-table table { min-width: 500px; }
        .drag-over { background-color: #e0f2fe; border-color: #0284c7; }
    </style>
</head>
<body class="bg-gradient-to-br from-slate-50 to-slate-100 min-h-screen py-8 px-4">

    <div class="max-w-6xl mx-auto">
        <!-- Header -->
        <div class="text-center mb-8">
            <h1 class="text-4xl font-bold text-slate-800 mb-2">
                <i class="fas fa-file-excel text-green-600 mr-2"></i>
                Excel Cleanup Tool
            </h1>
            <p class="text-slate-600">Upload your Excel file → Preview → Transform → Preview → Download</p>
        </div>

        <!-- Main card -->
        <div class="bg-white rounded-2xl shadow-xl overflow-hidden">
            <div class="p-6 border-b border-slate-200">
                <h2 class="text-xl font-semibold text-slate-700"><i class="fas fa-upload mr-2 text-blue-500"></i>1. Upload File</h2>
            </div>
            <div class="p-6">
                <!-- Drag & drop area -->
                <div id="dropzone" class="border-2 border-dashed border-slate-300 rounded-xl p-8 text-center cursor-pointer hover:border-blue-400 transition-all">
                    <i class="fas fa-cloud-upload-alt text-5xl text-slate-400 mb-3"></i>
                    <p class="text-slate-600 mb-2">Drag & drop your Excel file here, or <span class="text-blue-600 font-medium">click to browse</span></p>
                    <p class="text-sm text-slate-400">Supports .xlsx, .xls</p>
                    <input type="file" id="fileInput" accept=".xlsx,.xls" class="hidden" />
                </div>

                <!-- Upload preview section -->
                <div id="uploadPreview" class="mt-6 hidden">
                    <h3 class="font-medium text-slate-700 mb-2"><i class="fas fa-eye mr-1"></i> Uploaded file preview (first 10 rows)</h3>
                    <div id="uploadTable" class="preview-table border rounded-lg p-2 bg-slate-50"></div>
                </div>

                <!-- Process button (hidden until file uploaded) -->
                <div class="mt-6 text-center">
                    <button id="processBtn" disabled class="bg-blue-500 hover:bg-blue-600 text-white font-semibold py-2 px-6 rounded-lg shadow-md transition-all disabled:opacity-50 disabled:cursor-not-allowed">
                        <i class="fas fa-cogs mr-2"></i> Process File
                    </button>
                </div>
            </div>
        </div>

        <!-- Result card (hidden initially) -->
        <div id="resultCard" class="bg-white rounded-2xl shadow-xl overflow-hidden mt-8 hidden">
            <div class="p-6 border-b border-slate-200">
                <h2 class="text-xl font-semibold text-slate-700"><i class="fas fa-chalkboard-user mr-2 text-purple-500"></i>2. Transformed Preview</h2>
            </div>
            <div class="p-6">
                <div id="resultPreview" class="preview-table border rounded-lg p-2 bg-slate-50"></div>
                <div class="mt-6 text-center">
                    <button id="downloadBtn" class="bg-green-500 hover:bg-green-600 text-white font-semibold py-2 px-6 rounded-lg shadow-md transition-all">
                        <i class="fas fa-download mr-2"></i> Download Transformed Excel
                    </button>
                </div>
            </div>
        </div>

        <!-- Loading overlay -->
        <div id="loading" class="fixed inset-0 bg-black bg-opacity-50 flex items-center justify-center z-50 hidden">
            <div class="bg-white rounded-xl p-8 text-center">
                <i class="fas fa-spinner fa-spin text-4xl text-blue-500 mb-4"></i>
                <p class="text-slate-700">Processing your file... (cold start may take 30-60s)</p>
            </div>
        </div>
    </div>

    <script>
        // DOM elements
        const dropzone = document.getElementById('dropzone');
        const fileInput = document.getElementById('fileInput');
        const uploadPreviewDiv = document.getElementById('uploadPreview');
        const uploadTableDiv = document.getElementById('uploadTable');
        const processBtn = document.getElementById('processBtn');
        const resultCard = document.getElementById('resultCard');
        const resultPreviewDiv = document.getElementById('resultPreview');
        const downloadBtn = document.getElementById('downloadBtn');
        const loadingDiv = document.getElementById('loading');

        let uploadedFile = null;
        let processedBlob = null;

        // Helper: read file as array buffer
        function readFileAsBuffer(file) {
            return new Promise((resolve, reject) => {
                const reader = new FileReader();
                reader.onload = (e) => resolve(e.target.result);
                reader.onerror = reject;
                reader.readAsArrayBuffer(file);
            });
        }

        // Helper: parse workbook and render first N rows as HTML table
        function renderExcelPreview(arrayBuffer, targetDiv, rowsToShow = 10) {
            const workbook = XLSX.read(arrayBuffer, { type: 'array' });
            const firstSheet = workbook.Sheets[workbook.SheetNames[0]];
            const data = XLSX.utils.sheet_to_json(firstSheet, { header: 1, defval: "" });
            if (!data || data.length === 0) {
                targetDiv.innerHTML = '<p class="text-slate-500 italic">No data found.</p>';
                return;
            }
            const headers = data[0];
            const rows = data.slice(1, rowsToShow + 1);
            let html = '<table class="min-w-full text-sm">';
            // header row
            html += '<thead class="bg-slate-100"><tr>';
            headers.forEach(h => {
                html += `<th class="border px-2 py-1 text-left font-semibold">${escapeHtml(String(h))}</th>`;
            });
            html += '</tr></thead><tbody>';
            rows.forEach(row => {
                html += '<tr>';
                headers.forEach((_, idx) => {
                    let cell = row[idx] !== undefined ? row[idx] : "";
                    html += `<td class="border px-2 py-1">${escapeHtml(String(cell))}</td>`;
                });
                html += '</tr>';
            });
            html += '</tbody></table>';
            if (data.length > rowsToShow + 1) {
                html += `<p class="text-xs text-slate-400 mt-1">... and ${data.length - rowsToShow - 1} more rows</p>`;
            }
            targetDiv.innerHTML = html;
        }

        function escapeHtml(str) {
            return str.replace(/[&<>]/g, function(m) {
                if (m === '&') return '&amp;';
                if (m === '<') return '&lt;';
                if (m === '>') return '&gt;';
                return m;
            });
        }

        // Handle file selection
        function handleFile(file) {
            if (!file || !(file.name.endsWith('.xlsx') || file.name.endsWith('.xls'))) {
                alert('Please select a valid Excel file (.xlsx or .xls)');
                return false;
            }
            uploadedFile = file;
            // Show preview
            readFileAsBuffer(file).then(buffer => {
                renderExcelPreview(buffer, uploadTableDiv);
                uploadPreviewDiv.classList.remove('hidden');
                processBtn.disabled = false;
            }).catch(err => {
                console.error(err);
                alert('Error reading file for preview');
            });
            return true;
        }

        // Drag & drop events
        dropzone.addEventListener('dragover', (e) => {
            e.preventDefault();
            dropzone.classList.add('drag-over');
        });
        dropzone.addEventListener('dragleave', () => {
            dropzone.classList.remove('drag-over');
        });
        dropzone.addEventListener('drop', (e) => {
            e.preventDefault();
            dropzone.classList.remove('drag-over');
            const file = e.dataTransfer.files[0];
            handleFile(file);
        });
        dropzone.addEventListener('click', () => fileInput.click());
        fileInput.addEventListener('change', (e) => {
            if (e.target.files.length) handleFile(e.target.files[0]);
        });

        // Process button: send to /transform endpoint
        processBtn.addEventListener('click', async () => {
            if (!uploadedFile) return;
            loadingDiv.classList.remove('hidden');
            const formData = new FormData();
            formData.append('file', uploadedFile);
            try {
                const response = await fetch('/transform', { method: 'POST', body: formData });
                if (!response.ok) {
                    const errorText = await response.text();
                    throw new Error(errorText);
                }
                processedBlob = await response.blob();
                // Show preview of processed file
                const buffer = await processedBlob.arrayBuffer();
                renderExcelPreview(buffer, resultPreviewDiv);
                resultCard.classList.remove('hidden');
            } catch (err) {
                alert('Processing failed: ' + err.message);
                console.error(err);
            } finally {
                loadingDiv.classList.add('hidden');
            }
        });

        // Download button
        downloadBtn.addEventListener('click', () => {
            if (!processedBlob) return;
            const url = URL.createObjectURL(processedBlob);
            const a = document.createElement('a');
            a.href = url;
            a.download = 'transformed.xlsx';
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            URL.revokeObjectURL(url);
        });
    </script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route("/transform", methods=["POST"])
def transform():
    if "file" not in request.files:
        return "No file uploaded", 400
    file = request.files["file"]
    if file.filename == "":
        return "Empty filename", 400

    input_bytes = file.read()
    try:
        output_bytes = transform_in_memory(input_bytes)
        return send_file(
            io.BytesIO(output_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="transformed.xlsx"
        )
    except Exception as e:
        app.logger.exception("Transformation failed")
        return f"Error: {str(e)}", 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
